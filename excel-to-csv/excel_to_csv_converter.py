#! /usr/bin/env python

"""
A simple program that converts all excel files in working directory to CSV files
If an excel file has more than 1 sheet, each sheet will be it's own CSV file
"""

import os
import openpyxl
import csv
import shutil

PATH = '/home/ross/AllThingsPython/MyDev/atbs-scripts/excel-to-csv/excel_files/'

def xlsxToCsv(target_directory):
    os.chdir(target_directory)
    os.makedirs('Converted_Files', exist_ok=True)
    target_directory = 'Converted_Files/'
    print('Converting .xlsx files to .csv in the following directory:'
        '{target_directory}')
    try:
        for excel_file in os.listdir('.'):
            # Skip non-xlsx files, load the workbook object
            if excel_file.endswith('xlsx') and '~lock' not in excel_file:
                wb = openpyxl.load_workbook(excel_file)

                # Loop through every sheet in the workbook
                for sheet_name in wb.sheet_names:
                    sheet = wb[sheet_name] 

                    # Create CSV filename from the Excel filename & sheet title
                    csv_filename = excel_file[:-5] + '_' + sheet_name + '.csv'

                    # Create the csv.writer object for this CSV file
                    with open(csv_filename, 'w') as csvFile:
                        output_writer = csv.writer(csvFile)

                        range_data = []
                        for c in sheet.iter_cols(
                            sheet.min_column,
                            sheet.max_column,
                            sheet.min_row,
                            sheet.max_row
                            ):
                            column_data = []
                            for r in c:
                                column_data.append(r.value)
                            range_data.append(column_data)
                        for row in range_data:
                            output_writer.writerow(row)

                    # Print conversion message & move CSV to Converted_Files dir
                    print(f'File converted: \n{excel_file + '_' + sheet_name}\t'
                        '==>\t{csv_filename}\n')
                    shutil.move(csv_filename, target_directory)
        print('Conversion operations completed!\nExiting application. Goodbye!')

    except Exception as e:
        print('ERROR!\nError message: {}'.format(e))


if __name__ == '__main__':
    xlsxToCsv(PATH)