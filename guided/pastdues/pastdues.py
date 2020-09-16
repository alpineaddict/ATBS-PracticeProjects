#! /usr/bin/env python

# pastdues.py - check for past due balances within a spreadsheet and send email to corresponding user

import openpyxl, smtplib, sys, os

# open the spreadsheet and get the latest dues status
os.chdir('/home/ross/AllThingsPython/ATBS/PracticeProjects/Guided/pastdues')
wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb['Sheet1']
lastCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol).value

# Check each member's payment status
unpaidMembers = {}
for r in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email


# Log into email account
smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login('xxxcougarhunter69@gmail.com', 'ALTpass666!1')

# Send out reminder emails
for name, email in unpaidMembers.items():
    body = ("Subject: %s dues unpaid.\nDear %s,\nRecords show that you have not paid dues for %s. Please make"
            "this payment as soon as possible. Thank you!" % (latestMonth, name, latestMonth))
    print('Sending email to %s ...' % email)
    sendmailStatus = smtpObj.sendmail('xxxcougarhunter69@gmail.com', email, body)

    if sendmailStatus != {}:
        print('There was a problem sending email to %s: %s' % (email, sendmailStatus))
    smtpObj.quit()
